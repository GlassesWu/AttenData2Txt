from openpyxl import load_workbook
from tkinter import Tk
from re import match
from warnings import filterwarnings
from win32api import MessageBox
from win32con import MB_OK
from tkinter import filedialog

# 忽略警告
filterwarnings("ignore")

# 消除TK模块的小窗口
root = Tk()
root.withdraw()

# 文件路径提取（限制文件类型）
file_path = filedialog.askopenfilename(filetypes=[('XLSX','xlsx')])

# 未选择文件反馈
if file_path.strip() == '':
    """
    hWnd:0-独立弹框，不锁定原窗口
    IpText:框体文本
    IpCaption:弹框标题文本
    wType:弹框类型，MB_OK-只有一个确认键
    """
    MessageBox(0, '         未选择文件', '', MB_OK)

# 有选择文件
else:
    # 以异常处理方式执行流程
    try:
        # 获取sheet页,以list形式存储
        wb1 = load_workbook(file_path, data_only=True)
        sheets1 = wb1.sheetnames

        # 获取指定sheet页，源数据在excel第一sheet页且只有一页
        sheet1 = wb1[sheets1[0]]

        # 获取最大行数
        max_row = sheet1.max_row

        # 获取最大列数
        max_column = sheet1.max_column

        # 需要的列所在位置，zlist语句仅用于编程参考，无实际作用（考勤号码、日期、签到时间、签退时间）
        zlist = [2, 4, 8, 9, ]

        # 最终数据清单
        number_list = []

        # excel有效数据起始行：3
        for m in range(3, max_row+1):
            # 考勤号码转换为序列号，以8位数字格式存储
            serial_num = "%08d" % int(sheet1.cell(m, 2).value)
            # 正则日期对象，获取格式为'YYYY-MM-DD'，（\d···）为排序group
            date_obj = match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', str(sheet1.cell(m, 4).value))
            # 年
            year = date_obj.group(1)
            # 月
            mon = "%02d" % int(date_obj.group(2))
            # 日
            day = "%02d" % int(date_obj.group(3))
            # 日期
            date = year + mon + day
            # 签到时间
            clock_in = sheet1.cell(m, 8).value
            # 签退时间
            clock_out = sheet1.cell(m, 9).value

            # 无签到时间不处理
            if clock_in == None:
                pass

            # 有签到时间
            else:
                # 正则时间对象，获取格式为"HH:MM:SS"，先将时间内容转换为字符串，（\d···）为排序group
                time_obj = match(r'^(\d{2}):(\d{2}):(\d{2})$', str(clock_in))
                # 转换数据：P10-签到
                #          0001-考勤机终端编码，默认0001
                #          YYYYMMDD-年月日
                #          HH-时
                #          MM-分
                #          SS-秒
                #          YYYYMMDD-年月日
                #          HH-时
                #          MM-分
                #          SS-秒
                #          00000XXX-8位考勤机编码，X位数为实际编号长度
                convert_data = 'P100001' + date + str(time_obj.group(1)) + str(time_obj.group(2)) + str(time_obj.group(3)) + date + str(time_obj.group(1)) + \
                    str(time_obj.group(2)) + str(time_obj.group(3)) + serial_num
                # 转换数据放入最终数据清单
                number_list.append(convert_data)            

            # 无签退时间不处理
            if clock_out == None:
                pass

            # 有签退时间
            else:
                # 正则时间对象，获取格式为"HH:MM:SS"，先将时间内容转换为字符串，（\d···）为排序group
                time_obj = match(r'^(\d{2}):(\d{2}):(\d{2})$', str(clock_out))
                # 转换数据：P20-签退
                #          0001-考勤机终端编码，默认0001
                #          YYYYMMDD-年月日
                #          HH-时
                #          MM-分
                #          SS-秒
                #          YYYYMMDD-年月日
                #          HH-时
                #          MM-分
                #          SS-秒
                #          00000XXX-8位考勤机编码，X位数为实际编号长度
                convert_data = 'P200001' + date + str(time_obj.group(1)) + str(time_obj.group(2)) + str(time_obj.group(3)) + date + str(time_obj.group(1)) + \
                    str(time_obj.group(2)) + str(time_obj.group(3)) + serial_num
                # 转换数据放入最终数据清单
                number_list.append(convert_data)
        # 所有数据处理完毕后关闭excel
        wb1.close()

    # 异常处理
    except (TypeError,ValueError,AttributeError):
        # 统一反馈"文档内容无效"
        MessageBox(0, '         文档内容无效', '', MB_OK)

    # 无异常后续执行语句
    else:
        # 文件名：XXXX年XX月考勤转换数据.TXT，取最终数据清单第一条数据作参考
        filename = number_list[0][7:11] + '年' + number_list[0][11:13] + '月' + '考勤转换数据.txt'
        # 将最终数据清单按每条+换行形式写入文件
        file = open(filename, 'w')
        file.write('\n'.join(number_list))
        file.close()
        # 执行完毕反馈
        MessageBox(0, '           执行完毕', '', MB_OK)
